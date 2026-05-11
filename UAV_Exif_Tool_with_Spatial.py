#!/usr/bin/env python3 #Linux/macOS 下告诉系统：用 python3 执行这个文件,Windows 基本没用。
"""UAV Exif信息提取工具：照片 EXIF/XMP、视频 SRT、快照写 EXIF、空间索引查询。""" #这是程序说明

#导入python内置库
import csv
import os
import re
import struct
import sys
import threading
import time
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor

import cv2
import numpy as np
import piexif
import tkinter as tk
import tkinter.ttk as ttk
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PIL import Image
from rtree import index
from shapefile import Writer
from tkinter import filedialog, messagebox

#让import找到路径(把虚拟环境 site-packages 加入模块搜索路径)
sys.path.append(os.path.join(os.path.dirname(__file__), "..", ".venv", "Lib", "site-packages"))

GPS_INFO_TAG = 0x8825 #EXIF 里的 GPS 标签 ID
PHOTO_READ_SIZE = 65536 #读取 JPG 前 64KB。因为 EXIF 通常就在前面。这样比读取整个文件快。
JPG_EXTENSIONS = (".jpg", ".jpeg") #允许的照片格式。
APP1_XMP_MARKER = b"http://ns.adobe.com/" #XMP 标记。用于识别：APP1 段里是不是 XMP
#定义WGS84地理坐标系的所有参数,
#分别：地理坐标系，大地基准面，椭球体，赤道半径，扁率倒数，本初子午线，角度单位，1度对应的弧度值，权威代码，表明其标准代号为 EPSG:4326，是该坐标系的通用ID。
WGS84_PRT = 'GEOGCS["WGS 84",DATUM["WGS_1984",SPHEROID["WGS 84",6378137,298.257223563,AUTHORITY["EPSG","7030"]],AUTHORITY["EPSG","6326"]],PRIMEM["Greenwich",0,AUTHORITY["EPSG","8901"]],UNIT["degree",0.0174532925199433,AUTHORITY["EPSG","9122"]],AUTHORITY["EPSG","4326"]]'

#定义导出的列名。
FIELD_NAMES = [
    "文件名", "无人机偏航角", "无人机俯仰角", "无人机翻滚角",
    "云台偏航角", "云台俯仰角", "云台翻滚角",
    "GPS纬度", "GPS经度", "GPS高度"
]
#从原始数据中提取和处理姿态/坐标信息，并将其格式化写入Shp文件，同时确保坐标精度保留7位小数。
NUMERIC_FIELD_NAMES = FIELD_NAMES[1:]
SHAPE_RECORD_FIELDS = FIELD_NAMES[1:]
SHP_FIELDS = ["YAW", "PITCH", "ROLL", "GYAW", "GPITCH", "GROLL", "LAT", "LON", "ALT"]
COORD_FIELD_NAMES = FIELD_NAMES[7:9]
COORD_SHP_FIELDS = {"LAT", "LON"}
COORD_DECIMAL_PLACES = 7

#正则表达式模块：
#用于：从照片XMP元数据中提取飞行器和云台的6个姿态角
XMP_TAG_PATTERN = re.compile(
    r'(FlightYawDegree|FlightPitchDegree|FlightRollDegree|GimbalYawDegree|GimbalPitchDegree|GimbalRollDegree)="([^"]*)"'
)
#从视频SRT字幕文件中提取时间戳（时:分:秒,毫秒）
SRT_TIME_RE = re.compile(r"(\d+):(\d+):(\d+),(\d+) -->")
#从视频SRT字幕中提取经纬度和海拔（纬度、经度、相对高度）
SRT_COORD_RE = re.compile(
    r"\[latitude: ([\d.-]+)\].*?\[longitude: ([\d.-]+)\].*?\[rel_alt: ([\d.-]+)",
    re.S,
)
#从视频SRT字幕中提取云台的3个姿态角（gb_yaw, gb_pitch, gb_roll）
SRT_GB_RE = re.compile(r"\[gb_yaw: ([\d.-]+) gb_pitch: ([\d.-]+) gb_roll: ([\d.-]+)\]")
#将XMP标签名《映射》到固定索引位置，便于将提取的数据存入数组或列表
XMP_INDEX_MAP = {
    "FlightYawDegree": 0,
    "FlightPitchDegree": 1,
    "FlightRollDegree": 2,
    "GimbalYawDegree": 3,
    "GimbalPitchDegree": 4,
    "GimbalRollDegree": 5,
}

#主类：数据、函数、UI
#图形界面应用
class UAVExifTool:
    def __init__(self, root): #初始化函数。创建对象时自动执行。
        self.root = root #Tkinter 主窗口对象
        self.root.title("UAV Exif信息提取工具 - R-tree空间索引版")
        self.root.geometry("900x700")

        self.func_var = tk.StringVar(value="photo") #功能选择变量，默认处理"照片"（photo/video可选）
        self.query_mode = tk.StringVar(value="knn") #空间查询模式，默认"K近邻"（knn/radius可选）
        self.frame_count = tk.IntVar(value=40) #处理视频时的帧数，默认40帧
        self.thread_count = tk.IntVar(value=min(64, max(4, (os.cpu_count() or 8) * 2))) #线程数，根据CPU核心数自动计算
        self.progress_var = tk.DoubleVar(value=0) #进度条变量
        self.k_value = tk.IntVar(value=5) #K近邻查询的K值，默认5
        self.radius_value = tk.DoubleVar(value=0.001) #半径查询的半径值，默认0.001度（约111米）
        self.query_lon = tk.DoubleVar(value=112.946843) #查询点的经纬度海拔（默认长沙某地）
        self.query_lat = tk.DoubleVar(value=28.237419)
        self.query_alt = tk.DoubleVar(value=284.4)

        self.spatial_index = None #R-tree空间索引对象
        self.spatial_data = [] #存储所有空间数据记录
        self.spatial_coords = [] #存储所有坐标点（用于空间索引）
        self.index_built = False #标记空间索引是否已构建
        self.stop_event = threading.Event() #线程停止事件
        self.process_thread = None #处理线程对象

        self._create_widgets() #开始绘制界面

    def _create_widgets(self):
        notebook = ttk.Notebook(self.root) #创建标签页控件（类似浏览器的分页）
        notebook.pack(fill="both", expand=True, padx=10, pady=10)

        self.process_frame = tk.Frame(notebook) #创建“数据处理”页面的容器框架
        self.spatial_frame = tk.Frame(notebook) #创建“空间索引查询”页面的容器框架
        notebook.add(self.process_frame, text="数据处理") #将两个框架添加到标签页，并设置显示文字
        notebook.add(self.spatial_frame, text="空间索引查询")

        self._create_process_tab() #创建数据处理页的具体内容（文件选择、处理控制等）
        self._create_spatial_tab()

    def _create_process_tab(self):
        frame = self.process_frame #获取“数据处理”标签页的容器
#创建一个文本标签，显示“选择功能:”
#Tkinter 的网格布局。使用网格布局放置标签：• 第0行第0列• 四周留10像素间距• 靠左对齐（west）
        tk.Label(frame, text="选择功能:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
#创建功能选择下拉菜单，提供三种处理模式：• photo - 处理照片• video_snapshot - 视频快照（提取帧）• video_exif - 视频EXIF信息
        func_frame = tk.Frame(frame)
        func_frame.grid(row=0, column=1, columnspan=2, padx=10, pady=10, sticky="we")
        tk.OptionMenu(func_frame, self.func_var, "photo", "video_snapshot", "video_exif").pack(
            side="left", fill="x", expand=True
        )
#创建两行路径选择：1. 输入路径：选择待处理的照片/视频文件夹；2. 输出路径：选择结果保存位置
        self.input_path = self._path_row(frame, 1, "输入路径:", self._browse_input)
        self.output_path = self._path_row(frame, 2, "输出路径:", self._browse_output)
#数字调节框，设置视频处理时提取的帧数（30-50帧）
        tk.Label(frame, text="快照帧数(30-50):").grid(row=3, column=0, padx=10, pady=10, sticky="w")
        tk.Spinbox(frame, from_=30, to=50, textvariable=self.frame_count, width=5).grid(
            row=3, column=1, padx=10, sticky="w"
        )
#数字调节框，设置处理线程数（1-64个线程）
        tk.Label(frame, text="线程数:").grid(row=3, column=1, padx=10, pady=10, sticky="e")
        tk.Spinbox(frame, from_=1, to=64, textvariable=self.thread_count, width=5).grid(
            row=3, column=2, padx=10, sticky="w"
        )
#按钮容器框架，放置4个功能按钮
        btn_frame = tk.Frame(frame)
        btn_frame.grid(row=4, column=0, columnspan=3, padx=10, pady=10, sticky="we")
        self.start_button = self._pack_btn(btn_frame, "开始处理", self._start_process) #“开始处理”按钮，绑定_start_process方法
        self.stop_button = self._pack_btn(btn_frame, "停止", self._stop_process) #“停止”按钮，绑定_stop_process方法，初始为禁用状态
        self._pack_btn(btn_frame, "清空日志", self._clear_log) #“清空日志”按钮，绑定_clear_log方法
        self._pack_btn(btn_frame, "构建空间索引", self._build_spatial_index) #“构建空间索引”按钮，绑定_build_spatial_index方法
        self.stop_button.config(state=tk.DISABLED)
# 水平进度条，显示处理进度（0-100%）
        tk.Label(frame, text="进度:").grid(row=5, column=0, padx=10, pady=10, sticky="w")
        tk.Scale(
            frame, variable=self.progress_var, orient="horizontal", length=600, #进度条绑定的变量，控制进度显示
            from_=0, to=100, showvalue=False
        ).grid(row=5, column=1, columnspan=2, padx=10, pady=10, sticky="we")
#显示处理状态信息：1. 文件数统计2. 处理耗时3. 处理速度4. 线程数5. 程序状态6. 空间索引状态
        self.file_count_var = tk.StringVar(value="文件数: 0")
        self.time_var = tk.StringVar(value="耗时: 0.00秒")
        self.speed_var = tk.StringVar(value="速度: 0.00秒/文件")
        self.thread_var = tk.StringVar(value="线程数: 0")
        self.status_var = tk.StringVar(value="状态: 就绪")
        self.index_status_var = tk.StringVar(value="空间索引: 未构建")

        stats_frame = tk.Frame(frame) #创建状态栏，横向排列显示6个状态信息
        stats_frame.grid(row=6, column=0, columnspan=3, padx=10, pady=10, sticky="we")
        for var in (
            self.file_count_var, self.time_var, self.speed_var,
            self.thread_var, self.status_var, self.index_status_var
        ):
            tk.Label(stats_frame, textvariable=var).pack(side="left", padx=20) #将6个状态变量绑定到标签，实现动态更新
#创建"处理日志:"标签和多行日志显示区域
        tk.Label(frame, text="处理日志:").grid(row=7, column=0, padx=10, pady=10, sticky="nw")
        self.log_text = tk.Text(frame, height=15, wrap="word") #tk.Text和滚动条：：可滚动的文本框，高度15行，自动换行显示
        self.log_text.grid(row=7, column=1, columnspan=2, padx=10, pady=10, sticky="nsew")
        scrollbar = tk.Scrollbar(self.log_text, command=self.log_text.yview) #为日志文本框添加垂直滚动条
        scrollbar.pack(side="right", fill="y")
        self.log_text.config(yscrollcommand=scrollbar.set)
#自动填充输入/输出路径的默认值
        script_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(script_dir)
        self.input_path.insert(0, os.path.join(project_root, "第一次无人机数据"))
        self.output_path.insert(0, os.path.join(project_root, "输出结果"))
#设置网格布局权重，使日志区域可随窗口伸缩
        frame.grid_rowconfigure(7, weight=1)
        frame.grid_columnconfigure(1, weight=1)

    def _create_spatial_tab(self):
        frame = self.spatial_frame #选择查询方式：• K最近邻：查询最近的K个点• 范围查询：查询指定半径内的点
        tk.Label(frame, text="查询模式:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        tk.Radiobutton(frame, text="K最近邻", variable=self.query_mode, value="knn").grid(
            row=0, column=1, padx=10, sticky="w"
        )
        tk.Radiobutton(frame, text="范围查询", variable=self.query_mode, value="range").grid(
            row=0, column=2, padx=10, sticky="w"
        )
#设置查询参数：1. K值（K近邻查询）2. 半径（范围查询，单位：度）3. 查询点的经度、纬度、高度
        self._entry_row(frame, 1, 0, "K值:", self.k_value)
        self._entry_row(frame, 1, 1, "半径(度):", self.radius_value)
        self._entry_row(frame, 2, 0, "查询经度:", self.query_lon)
        self._entry_row(frame, 2, 1, "查询纬度:", self.query_lat)
        self._entry_row(frame, 3, 0, "查询高度:", self.query_alt)
#触发查询操作，调用_execute_spatial_query方法
        tk.Button(frame, text="执行查询", command=self._execute_spatial_query).grid(
            row=4, column=0, columnspan=3, padx=10, pady=10, sticky="we"
        )
        tk.Label(frame, text="查询结果:").grid(row=5, column=0, padx=10, pady=10, sticky="nw")
#显示查询结果，高度25行，带垂直滚动条
        self.query_result_text = tk.Text(frame, height=25, wrap="word")
        self.query_result_text.grid(row=5, column=1, columnspan=2, padx=10, pady=10, sticky="nsew")
        scrollbar = tk.Scrollbar(self.query_result_text, command=self.query_result_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.query_result_text.config(yscrollcommand=scrollbar.set)
#使结果文本框区域可随窗口大小变化而伸缩
        frame.grid_rowconfigure(5, weight=1)
        frame.grid_columnconfigure(1, weight=1)
#创建路径选择行（标签+输入框+浏览按钮），返回输入框对象
    def _path_row(self, frame, row, label, command):
        tk.Label(frame, text=label).grid(row=row, column=0, padx=10, pady=10, sticky="w")
        entry = tk.Entry(frame, width=50); entry.grid(row=row, column=1, padx=10, pady=10, sticky="we")
        tk.Button(frame, text="浏览", command=command).grid(row=row, column=2, padx=10, pady=10); return entry
#创建参数输入行（标签+输入框），按两列为一组排列
    def _entry_row(self, frame, row, col_group, label, variable):
        col = col_group * 2
        tk.Label(frame, text=label).grid(row=row, column=col, padx=10, pady=10, sticky="w")
        tk.Entry(frame, textvariable=variable, width=15).grid(row=row, column=col + 1, padx=10, sticky="w")
#创建按钮并水平排列，按钮可随容器扩展
    def _pack_btn(self, frame, text, command):
        btn = tk.Button(frame, text=text, command=command); btn.pack(side="left", padx=10, fill="x", expand=True); return btn
#打开文件夹选择对话框，将选择的路径填入输入路径框
    def _browse_input(self):
        path = filedialog.askdirectory()
        if path:
            self.input_path.delete(0, tk.END)
            self.input_path.insert(0, path)
#打开文件夹选择对话框，将选择的路径填入输出路径框
    def _browse_output(self):
        path = filedialog.askdirectory()
        if path:
            self.output_path.delete(0, tk.END)
            self.output_path.insert(0, path)

    def _start_process(self):
        self.start_button.config(state=tk.DISABLED) #禁用"开始处理"按钮，启用"停止"按钮，防止重复启动
        self.stop_button.config(state=tk.NORMAL)
        self.stop_event.clear() #清除停止标志，允许新任务运行
        self._clear_log() #清除之前的日志记录，开始新的日志
        self.progress_var.set(0) #将进度条归零，准备重新显示进度
        self.file_count_var.set("文件数: 0")
        self.time_var.set("耗时: 0.00秒")
        self.speed_var.set("速度: 0.00秒/文件")
        self.thread_var.set(f"线程数: {self.thread_count.get()}")#显示当前设置的线程数量
        self.status_var.set("状态: 处理中") #更新状态
        self.process_thread = threading.Thread(target=self._process_files, daemon=True)
        self.process_thread.start()
#停止后台处理线程
    def _stop_process(self):
        self.stop_event.set()
        self.status_var.set("状态: 已停止")
#清空日志显示区域
    def _clear_log(self):
        self.log_text.delete(1.0, tk.END)
#添加带时间戳的日志消息，并自动滚动到底部
    def _log(self, message):
        self.log_text.insert(tk.END, f"{time.strftime('%H:%M:%S')} - {message}\n")
        self.log_text.see(tk.END)
        self.root.update()
#更新进度条，显示处理进度百分比
    def _update_progress(self, current, total):
        self.progress_var.set(0 if total == 0 else current / total * 100)
        self.root.update()
#处理完成后恢复按钮状态
    def _set_done_state(self):
        self.start_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
#从界面获取输入路径、输出路径、功能模式和线程数
    def _process_files(self):
        input_dir = self.input_path.get().strip()
        output_dir = self.output_path.get().strip()
        func = self.func_var.get()
        thread_count = self.thread_count.get()
#检查输入路径是否存在，不存在则报错并返回
        if not os.path.exists(input_dir):
            self._log(f"错误: 输入路径不存在 {input_dir}")
            self._set_done_state()
            return
#确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)
        start_time = time.time()
        results, read_time = { #根据选择的功能调用对应的处理方法：1. photo→ 处理照片2. video_snapshot→ 处理视频快照3. video_exif→ 处理视频EXIF
            "photo": lambda: self._process_photos(input_dir, thread_count),
            "video_snapshot": lambda: self._process_video_snapshot(input_dir, output_dir),
            "video_exif": lambda: self._process_video_exif(input_dir),
        }[func]()
#计算处理时间、文件数量、平均处理速度，评估性能是否达标
        total_time = time.time() - start_time
        if not results:
            self.status_var.set("状态: 无数据")
            self._log("处理完成，无可输出数据")
            self._set_done_state()
            return
#调用_save_results将处理结果保存到文件
        self.spatial_data = results
        count = len(results)
        avg = read_time / count
        self.file_count_var.set(f"文件数: {count}")
        self.time_var.set(f"耗时: {total_time:.2f}秒")
        self.speed_var.set(f"读取速度: {avg:.6f}秒/文件")
        self.status_var.set("状态: 达标" if avg < 0.001 else "状态: 未达标")
#更新所有状态变量，恢复界面状态
        save_start = time.time()
        self._save_results(results, output_dir, func)
        self._log(f"保存耗时: {time.time() - save_start:.2f}秒")
        self._log("处理完成")
        self._update_progress(1, 1)
        self._set_done_state()

    def _process_photos(self, input_dir, thread_count): ##定义函数。参数：input_dir输入照片目录；；；thread_count用户设置的线程数
        photo_items = sorted( #扫描所有JPG文件，并行读取 EXIF，最后返回结果列表
            #sorted()表示排序，因为：os.scandir()返回顺序不稳定，所以这里强制排序。
            (entry.name, entry.path) #生成器表达式
            for entry in os.scandir(input_dir) #os作用：扫描目录例如：IMG001.jpg，IMG002.jpg，test.txt；它返回：DirEntry对象
            if entry.is_file() and entry.name.lower().endswith(JPG_EXTENSIONS)
        )
        if not photo_items:
            return [], 0
        #智能确定线程数
        workers = self._resolve_photo_workers(thread_count, len(photo_items))
        #记录日志
        self._log(f"发现 {len(photo_items)} 张照片，photo 实际使用 {workers} 个线程")
        #并发读取EXIF（使用线程池）
        read_start = time.perf_counter()# 高精度计时
        with ThreadPoolExecutor(max_workers=workers) as executor:
            results = list(executor.map(self._read_photo_record, photo_items))
        read_time = time.perf_counter() - read_start
        self._update_progress(len(results), len(photo_items)) #更新进度并返回
        return results, read_time
#智能线程数计算
    def _resolve_photo_workers(self, requested_workers, file_count):
        cpu_count = os.cpu_count() or 8 # 获取CPU核心数，默认8
        stable_cap = max(4, min(16, cpu_count * 2)) # 稳定范围：4-16，不超过CPU*2
        return max(1, min(requested_workers, file_count, stable_cap))

    def _read_photo_record(self, photo_item):
        file_name, file_path = photo_item
        row = self._read_exif_binary_fast(file_path) ## 快速读取EXIF
        row["文件名"] = file_name # 添加文件名到结果
        return row
#快速读取照片的EXIF二进制数据
    def _read_exif_binary_fast(self, file_path): #定义方法，接收文件路径。
        result = [0.0] * len(NUMERIC_FIELD_NAMES) #初始化结果列表，长度与NUMERIC_FIELD_NAMES相同，初始值为0.0。
        with open(file_path, "rb") as f: #以二进制方式打开文件，读取前PHOTO_READ_SIZE个字节
            data = f.read(PHOTO_READ_SIZE)
        if len(data) < 2 or data[:2] != b"\xff\xd8": #检查数据长度是否至少2字节，且前两个字节是否为JPEG文件的起始标记（0xFFD8），如果不是，返回空结果字典。
            return self._result_list_to_dict(result)

        offset = 2 #设置偏移量，从2开始（跳过起始标记）
        while offset + 4 <= len(data):
            if data[offset] != 0xFF: #如果当前偏移的字节不是0xFF（标记开始），则跳出循环。
                break
            length = (data[offset + 2] << 8) | data[offset + 3] #计算当前段的长度（位于偏移+2和偏移+3的两个字节），如果长度小于2，跳出循环。
            if length < 2:
                break
            next_offset = offset + 2 + length #计算下一个段的偏移，如果超出数据长度，跳出循环。
            if next_offset > len(data):
                break
#如果当前段标记是0xE1（APP1段，通常包含EXIF和XMP数据），则提取该段数据。如果前6个字节是"Exif\x00\x00"，则调用parse_exif_data解析EXIF数据；如果APP1_XMP_MARKER（可能是"http://ns.adobe.com/xap/1.0/"）在该段数据中，则调用parse_xmp_fast解析XMP数据。
            if data[offset + 1] == 0xE1:
                app1_data = data[offset + 4:next_offset]
                if app1_data[:6] == b"Exif\x00\x00":
                    self._parse_exif_data(app1_data[6:], result)
                if APP1_XMP_MARKER in app1_data:
                    self._parse_xmp_fast(app1_data.decode("utf-8", errors="ignore"), result)
            offset = next_offset #将偏移移动到下一个段，继续循环。
        return self._result_list_to_dict(result) #将结果列表转换为字典并返回。

    @staticmethod #定义静态方法，接收exif数据和结果列表。
    def _parse_exif_data(exif_data, result): #如果exif_data长度小于10，返回。
        if len(exif_data) < 10:
            return
        byte_order = exif_data[:2] #获取字节序（II表示小端，MM表示大端），如果不是这两种，返回。
        if byte_order not in (b"II", b"MM"):
            return
        fmt = "<" if byte_order == b"II" else ">" #根据字节序设置格式化字符（<表示小端，>表示大端）
        if struct.unpack_from(fmt + "H", exif_data, 2)[0] != 0x002A: #检查从偏移2开始的2个字节是否等于0x002A（EXIF标准标记），如果不是，返回。
            return
        #从偏移4读取4个字节（无符号长整型）作为第一个IFD的偏移，并调用_find_gps_info方法。
        UAVExifTool._find_gps_info(exif_data, struct.unpack_from(fmt + "L", exif_data, 4)[0], fmt, result)

    @staticmethod #定义静态方法，接收exif数据、偏移、格式和结果列表。
    def _find_gps_info(exif_data, offset, fmt, result): #检查偏移是否有效，无效则返回。
        if offset < 0 or offset + 2 > len(exif_data):
            return
        num_tags = struct.unpack_from(fmt + "H", exif_data, offset)[0] #从偏移处读取2个字节，得到标签数量
        entry_base = offset + 2#计算第一个标签的基地址（偏移+2）。
        for tag_index in range(num_tags):
            tag_offset = entry_base + tag_index * 12
            if tag_offset + 12 > len(exif_data):
                return
            if struct.unpack_from(fmt + "H", exif_data, tag_offset)[0] == GPS_INFO_TAG: #检查当前标签偏移是否超出数据范围，是则返回。
                gps_offset = struct.unpack_from(fmt + "L", exif_data, tag_offset + 8)[0]
                UAVExifTool._parse_gps_info(exif_data, gps_offset, fmt, result) #读取当前标签的ID（2字节），如果等于GPS_INFO_TAG（0x8825），则从标签偏移+8处读取4字节作为GPS信息的偏移，并调用_parse_gps_info方法，然后返回。
                return

    @staticmethod #定义静态方法，接收exif数据、偏移、格式和结果列表。
    def _parse_gps_info(exif_data, offset, fmt, result): #检查偏移是否有效，无效则返回。
        if offset < 0 or offset + 2 > len(exif_data):
            return
        num_tags = struct.unpack_from(fmt + "H", exif_data, offset)[0] #读取GPS标签数量。
        entry_base = offset + 2 #计算第一个GPS标签的基地址。
        gps_values = {} #初始化gps_values字典，用于存储解析的GPS值。

        for tag_index in range(num_tags): #检查标签偏移是否有效，无效则跳出循环。
            tag_offset = entry_base + tag_index * 12# 计算当前标签的偏移位置（每个标签占12字节）
            if tag_offset + 12 > len(exif_data): # 检查偏移是否越界，越界则终止循环
                break
            # 解包标签ID、格式（忽略）和值计数
            tag_id, _, count = struct.unpack_from(fmt + "HHL", exif_data, tag_offset)
            value_offset = struct.unpack_from(fmt + "L", exif_data, tag_offset + 8)[0]    # 解包实际值的存储偏移量
            inline_value = exif_data[tag_offset + 8:tag_offset + 12] #提取标签内联值（标签最后4个字节，有时值直接存储在这里）
            # 处理纬度/经度参考（N/S/E/W）
            if tag_id in (1, 3):
                gps_values[tag_id] = inline_value[:count].decode("ascii", errors="ignore").strip("\x00 ") or ("N" if tag_id == 1 else "E")   # 解码内联值，清理空字符，默认N（纬度）或E（经度）
            elif tag_id in (2, 4):
                gps_values[tag_id] = UAVExifTool._parse_gps_coordinate(exif_data, value_offset, fmt)
            elif tag_id == 6:  # 处理海拔高度
                gps_values[tag_id] = UAVExifTool._parse_gps_altitude(exif_data, value_offset, fmt)
        # 从字典中获取经纬度和高度
        lat, lon, alt = gps_values.get(2), gps_values.get(4), gps_values.get(6)
        lat_ref, lon_ref = gps_values.get(1, "N"), gps_values.get(3, "E")# 获取经纬度参考方向，默认N和E
        if lat is not None and lon is not None:
            result[6] = round(-lat if lat_ref == "S" else lat, 9)# 若经纬度有效则处理符号并保留9位小数
            result[7] = round(-lon if lon_ref == "W" else lon, 9)
        if alt is not None:# 若高度有效则保留9位小数
            result[8] = round(alt, 9)

    @staticmethod
    def _parse_gps_coordinate(exif_data, offset, fmt):
        if offset < 0 or offset + 24 > len(exif_data):# 检查偏移是否有效，无效返回0.0
            return 0.0
        # 解包度、分、秒的分子和分母（各占8字节）
        deg_num, deg_den, min_num, min_den, sec_num, sec_den = struct.unpack_from(fmt + "LLLLLL", exif_data, offset)
        return (
            (deg_num / deg_den if deg_den else 0.0) # 转换为十进制度数：度 + 分/60 + 秒/3600
            + (min_num / min_den / 60 if min_den else 0.0)
            + (sec_num / sec_den / 3600 if sec_den else 0.0)
        )

    @staticmethod
    def _parse_gps_altitude(exif_data, offset, fmt): # 检查偏移是否有效，无效返回0.0
        if offset < 0 or offset + 8 > len(exif_data):
            return 0.0
        alt_num, alt_den = struct.unpack_from(fmt + "LL", exif_data, offset)  # 解包高度的分子和分母
        return alt_num / alt_den if alt_den else 0.0    # 计算高度值，分母为0则返回0.0

    @staticmethod
    def _parse_xmp_fast(xmp_segment, result):    # 用正则提取XMP中的标签和值
        for tag_name, tag_value in XMP_TAG_PATTERN.findall(xmp_segment):
            if tag_name in XMP_INDEX_MAP:     # 若标签在映射表中则转换为浮点数并保留9位小数
                result[XMP_INDEX_MAP[tag_name]] = round(float(tag_value), 9)

    @staticmethod
    def _result_list_to_dict(result, file_name=""): # 初始化行字典，包含文件名
        row = {"文件名": file_name}
        row.update(zip(NUMERIC_FIELD_NAMES, result))  # 将数值字段名与结果列表合并
        return row

    def _process_video_snapshot(self, input_dir, output_dir):   # 获取目录下所有.srt文件
        srt_files = [f for f in os.listdir(input_dir) if f.lower().endswith(".srt")]
        if not srt_files:  # 无SRT文件则返回空结果
            return [], 0

        results = []   # 记录开始处理时间
        read_start = time.perf_counter()
        for index_srt, srt_file in enumerate(srt_files, 1):   # 遍历SRT文件，索引从1开始
            if self.stop_event.is_set():  # 若收到停止信号则中断
                break
            srt_path = os.path.join(input_dir, srt_file)       # 构建SRT和对应视频的完整路径
            video_path = os.path.join(input_dir, os.path.splitext(srt_file)[0] + ".MP4")
            self._log(f"处理视频: {os.path.basename(video_path)}")   # 记录正在处理的视频

            gps_data = self._parse_srt_file(srt_path)    # 解析SRT文件获取GPS数据
            if not gps_data:   # 无GPS数据则跳过
                continue
            self._log(f"解析到 {len(gps_data)} 条GPS数据")    # 记录解析到的GPS数据量
            # 打开视频文件
            cap = cv2.VideoCapture(video_path)
            if not cap.isOpened():    # 打开失败则跳过
                continue
            total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))   # 获取视频总帧数
            fps = cap.get(cv2.CAP_PROP_FPS)     # 获取视频帧率
            interval = max(1, len(gps_data) // self.frame_count.get())     # 计算截图间隔（至少为1）
            # 遍历排序后的GPS帧ID
            for i, fid in enumerate(sorted(gps_data)):
                if self.stop_event.is_set():     # 若收到停止信号则中断
                    break
                if i % interval:     # 按间隔抽样
                    continue
                gps_info = gps_data[fid]     # 获取当前帧的GPS信息
                frame_idx = max(0, min(int(gps_info["timestamp"] * fps), total_frames - 1))    # 计算目标帧索引，确保在有效范围内
                cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)  # 跳转到目标帧
                ok, frame = cap.read()   # 读取帧图像
                if not ok:     # 读取失败则跳过
                    continue
                # 生成截图文件名
                snap_name = f"{os.path.splitext(os.path.basename(video_path))[0]}_snap{i:03d}.jpg"
                snap_path = os.path.join(output_dir, snap_name)        # 构建截图保存路径
                Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)).save(snap_path, "JPEG")      # 将BGR转为RGB并保存为JPEG
                self._write_jpg_exif(snap_path, gps_info)     # 向截图写入GPS Exif信息
                results.append(self._gps_info_to_row(snap_name, gps_info))     # 将结果转为字典并添加到列表
            cap.release()   # 释放视频资源
            self._update_progress(index_srt, len(srt_files))    # 更新处理进度
        return results, time.perf_counter() - read_start    # 返回结果列表和总耗时

    def _write_jpg_exif(self, jpg_path, gps_info):    # 从GPS信息中提取经纬度和高度
        lat, lon, alt = gps_info["latitude"], gps_info["longitude"], gps_info["altitude"]
        if lat == 0 and lon == 0:    # 若经纬度均为0则不处理
            return

        # 定义十进制度转度分秒的内部函数
        def dms(value):
            abs_value = abs(value)
            degrees = int(abs_value)
            minutes_full = (abs_value - degrees) * 60
            minutes = int(minutes_full)
            seconds = int((minutes_full - minutes) * 60 * 10000)
            return (degrees, 1), (minutes, 1), (seconds, 10000)

        # 构建GPS IFD字典
        gps_ifd = {
            piexif.GPSIFD.GPSLatitudeRef: b"N" if lat >= 0 else b"S",
            piexif.GPSIFD.GPSLatitude: dms(lat),
            piexif.GPSIFD.GPSLongitudeRef: b"E" if lon >= 0 else b"W",
            piexif.GPSIFD.GPSLongitude: dms(lon),
            piexif.GPSIFD.GPSAltitudeRef: 0 if alt >= 0 else 1,
            piexif.GPSIFD.GPSAltitude: (int(abs(alt) * 10000), 10000),
        }
        # 生成Exif字节流（仅含GPS）
        exif_bytes = piexif.dump({"0th": {}, "Exif": {}, "GPS": gps_ifd, "1st": {}, "thumbnail": None})
        Image.open(jpg_path).save(jpg_path, "JPEG", exif=exif_bytes)    # 打开图片并重新保存以写入Exif

    def _parse_srt_file(self, srt_path):
        gps_data = {}
        with open(srt_path, "r", encoding="utf-8") as f:   # 打开并读取SRT文件内容
            content = f.read()

        for entry in re.split(r"\n\n+", content):    # 按空行分割每个条目
            lines = entry.strip().split("\n")    # 去除空白并按行分割
            if len(lines) < 3 or not lines[0].strip().isdigit():      # 至少3行且首行为数字
                continue
            time_match = SRT_TIME_RE.match(lines[1])        # 匹配时间码行
            if not time_match:        # 匹配失败则跳过
                continue
            # 查找包含经纬度的行
            gps_line = next((line for line in lines[2:] if "latitude" in line and "longitude" in line), "")
            coord_match = SRT_COORD_RE.search(gps_line)     # 匹配坐标值
            if not coord_match:      # 匹配失败则跳过
                continue
                # 提取帧ID
            fid = int(lines[0].strip())
            hh, mm, ss, ms = map(int, time_match.groups())     # 提取时分秒毫秒
            gb_match = SRT_GB_RE.search(gps_line)    # 匹配云台角度
            gyaw, gpitch, groll = (map(float, gb_match.groups()) if gb_match else (0.0, 0.0, 0.0))      # 提取云台角度，无则为0
            # 存入GPS数据字典
            gps_data[fid] = {
                "fid": fid,
                "timestamp": hh * 3600 + mm * 60 + ss + ms / 1000,
                "latitude": float(coord_match.group(1)),
                "longitude": float(coord_match.group(2)),
                "altitude": float(coord_match.group(3)),
                "yaw": 0.0,
                "pitch": 0.0,
                "roll": 0.0,
                "gimbal_yaw": float(gyaw),
                "gimbal_pitch": float(gpitch),
                "gimbal_roll": float(groll),
            }
        return gps_data    # 返回解析后的GPS数据

    def _process_video_exif(self, input_dir):
        srt_files = [f for f in os.listdir(input_dir) if f.lower().endswith(".srt")]    # 获取目录下所有.srt文件
        if not srt_files:   # 无SRT文件则返回空结果
            return [], 0

        results = []  # 记录开始处理时间
        read_start = time.perf_counter()
        for index_srt, srt_file in enumerate(srt_files, 1):   # 遍历SRT文件
            if self.stop_event.is_set():        # 若收到停止信号则中断
                break
            video_file = os.path.splitext(srt_file)[0] + ".MP4"        # 生成对应视频文件名
            self._log(f"处理SRT文件: {srt_file}")      # 记录正在处理的SRT
            gps_data = self._parse_srt_file(os.path.join(input_dir, srt_file))     # 解析SRT获取GPS数据
            for fid, gps_info in gps_data.items():      # 遍历每条GPS数据
                results.append(self._gps_info_to_row(f"{video_file}_FID{fid:03d}", gps_info))     # 生成带FID的文件名并转为行字典
            self._log(f"已提取 {len(gps_data)} 条GPS记录")     # 记录提取的GPS记录数
            self._update_progress(index_srt, len(srt_files))
        return results, time.perf_counter() - read_start  # 返回结果列表和总耗时

    def _gps_info_to_row(self, file_name, gps_info):
        return {    # 将GPS信息转为标准化字典行
            "文件名": file_name,
            "无人机偏航角": gps_info.get("yaw", 0.0),
            "无人机俯仰角": gps_info.get("pitch", 0.0),
            "无人机翻滚角": gps_info.get("roll", 0.0),
            "云台偏航角": gps_info.get("gimbal_yaw", 0.0),
            "云台俯仰角": gps_info.get("gimbal_pitch", 0.0),
            "云台翻滚角": gps_info.get("gimbal_roll", 0.0),
            "GPS纬度": gps_info.get("latitude", 0.0),
            "GPS经度": gps_info.get("longitude", 0.0),
            "GPS高度": gps_info.get("altitude", 0.0),
        }

    @staticmethod
    def _is_coordinate_field(field_name):   # 判断字段是否为坐标类字段
        return field_name in COORD_FIELD_NAMES

    @staticmethod
    def _format_coordinate_value(value):    # 将坐标值格式化为指定位数的小数字符串
        return f"{float(value):.{COORD_DECIMAL_PLACES}f}"

    @classmethod
    def _normalize_export_row(cls, row):  # 复制行字典
        normalized = dict(row)
        for field_name in COORD_FIELD_NAMES:   # 遍历所有坐标字段
            value = normalized.get(field_name)
            if value not in (None, ""):      # 若值非空则格式化
                normalized[field_name] = cls._format_coordinate_value(value)
        return normalized   # 返回标准化后的行

    @classmethod
    def _normalize_shapefile_row(cls, row):  # 复制行字典
        normalized = dict(row)
        for field_name in COORD_FIELD_NAMES:    # 遍历所有坐标字段
            value = normalized.get(field_name)
            if value not in (None, ""):    # 若值非空则四舍五入到指定位数
                normalized[field_name] = round(float(value), COORD_DECIMAL_PLACES)
        return normalized   # 返回标准化后的行

    def _save_results(self, results, output_dir, func):
        prefix = "照片Exif信息" if func == "photo" else "视频Exif信息"    # 根据功能类型确定文件名前缀
        base_path = os.path.join(output_dir, prefix)   # 构建输出文件基础路径
        export_rows = [self._normalize_export_row(row) for row in results]    # 标准化所有导出行

        with open(base_path + ".csv", "w", newline="", encoding="utf-8-sig") as f:    # 保存为CSV文件
            writer = csv.DictWriter(f, fieldnames=FIELD_NAMES)
            writer.writeheader()
            writer.writerows(export_rows)
        # 保存为XML文件
        root = ET.Element("ExifData")
        for row in export_rows:
            item = ET.SubElement(root, "Item")
            for field_name in FIELD_NAMES:
                ET.SubElement(item, field_name).text = str(row.get(field_name, ""))
        ET.ElementTree(root).write(base_path + ".xml", encoding="utf-8", xml_declaration=True)
        # 保存为Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = "Exif数据"
        ws.append(FIELD_NAMES)
        coord_columns = [    # 确定坐标列的位置（从1开始）
            index + 1 for index, field_name in enumerate(FIELD_NAMES) if self._is_coordinate_field(field_name)
        ]
        for row in results:   # 写入数据行并设置坐标列格式
            ws.append([row.get(field, "") for field in FIELD_NAMES])
            current_row = ws.max_row
            for column_index in coord_columns:   # 调整坐标列宽度
                ws.cell(row=current_row, column=column_index).number_format = f"0.{('0' * COORD_DECIMAL_PLACES)}"
        for column_index in coord_columns:
            ws.column_dimensions[get_column_letter(column_index)].width = 14
        wb.save(base_path + ".xlsx")

        self._write_shapefile_pyshp(results, base_path)   # 保存为Shapefile点文件

    def _write_shapefile_pyshp(self, results, output_path):
        valid_results = [r for r in results if r["GPS经度"] != 0 and r["GPS纬度"] != 0]    # 筛选经纬度非零的有效结果
        if not valid_results:    # 无有效结果则返回
            return
        # 创建点类型Shapefile写入器
        writer = Writer(output_path, shapeType=1)
        for field_name in SHP_FIELDS:   # 添加字段：坐标类字段保留指定位小数，其他保留9位
            writer.field(field_name, "N", 12, COORD_DECIMAL_PLACES if field_name in COORD_SHP_FIELDS else 9)
        for row in valid_results:   # 写入每个点和属性
            normalized_row = self._normalize_shapefile_row(row)
            writer.point(normalized_row["GPS经度"], normalized_row["GPS纬度"])
            writer.record(*[normalized_row.get(field, 0) for field in SHAPE_RECORD_FIELDS])
        writer.close()
        # 写入WGS84投影文件
        with open(output_path + ".prj", "w") as f:
            f.write(WGS84_PRT)
        self._log(f"已保存 {output_path}.shp")   # 记录保存日志
        self._write_shapefile_line(valid_results, output_path)    # 同时生成轨迹线文件

    def _write_shapefile_line(self, valid_results, output_path):
        if len(valid_results) < 2:   # 点数量少于2无法成线则返回
            return
        # 提取所有点的经纬度坐标
        points = [[r["GPS经度"], r["GPS纬度"]] for r in valid_results]
        length = sum(    # 计算轨迹总长度（欧氏距离累加）
            np.sqrt((points[i][0] - points[i - 1][0]) ** 2 + (points[i][1] - points[i - 1][1]) ** 2)
            for i in range(1, len(points))
        )
        # 创建线类型Shapefile写入器
        line_path = output_path + "_line"
        writer = Writer(line_path, shapeType=3)
        writer.field("NAME", "C", 50)   # 添加属性字段
        writer.field("POINTS", "N", 10)
        writer.field("LENGTH", "N", 15, 6)
        writer.line([points])   # 写入轨迹线
        writer.record("Flight Path", len(points), length)   # 写入属性记录
        writer.close()
        # 写入WGS84投影文件
        with open(line_path + ".prj", "w") as f:
            f.write(WGS84_PRT)

    def _build_spatial_index(self):
        if not self.spatial_data:   # 若无空间数据则警告并返回
            messagebox.showwarning("警告", "没有可用数据！请先处理数据。")
            return
        # 提取有效经纬度和高度坐标
        self.spatial_coords = [
            (row["GPS经度"], row["GPS纬度"], row["GPS高度"])
            for row in self.spatial_data
            if row.get("GPS纬度") and row.get("GPS经度")
        ]
        if not self.spatial_coords:   # 若无有效坐标则警告并返回
            messagebox.showwarning("警告", "没有有效的GPS坐标数据！")
            return
        # 记录索引构建开始时间
        start_time = time.perf_counter()
        self.spatial_index = index.Index()   # 创建RTree空间索引
        for i, (lon, lat, _) in enumerate(self.spatial_coords):   # 向索引插入每个点（用经纬度作为边界框）
            self.spatial_index.insert(i, (lon, lat, lon, lat))
        build_time = (time.perf_counter() - start_time) * 1000   # 计算构建耗时（毫秒）
        # 标记索引已构建
        self.index_built = True
        self.index_status_var.set(f"空间索引: 已构建 ({len(self.spatial_coords)}个点, {build_time:.2f}ms)")  # 更新界面状态显示
        self._log(f"空间索引构建完成！点数量: {len(self.spatial_coords)}, 耗时: {build_time:.2f}ms")    # 记录日志

    def _execute_spatial_query(self):
        if not self.index_built or self.spatial_index is None:    # 若索引未构建则提示错误
            self.query_result_text.insert(tk.END, "错误: 请先构建空间索引！\n")
            return
        # 获取查询点的经纬度、高度和查询模式
        lon, lat, alt = self.query_lon.get(), self.query_lat.get(), self.query_alt.get()
        mode = self.query_mode.get()
        start_time = time.perf_counter()   # 记录查询开始时间
        header = (   # 构建查询结果头部信息
            f"{'=' * 60}\n"
            f"查询时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"查询模式: {'K最近邻' if mode == 'knn' else '范围查询'}\n"
            f"查询点: 经度={lon:.6f}, 纬度={lat:.6f}, 高度={alt:.3f}\n"
            f"{'=' * 60}\n\n"
        )
        # 执行查询：K近邻或范围查询
        ids = (
            list(self.spatial_index.nearest((lon, lat, lon, lat), self.k_value.get()))
            if mode == "knn"
            else list(self.spatial_index.intersection((
                lon - self.radius_value.get(), lat - self.radius_value.get(),
                lon + self.radius_value.get(), lat + self.radius_value.get()
            )))
        )
        # 计算查询耗时（毫秒）
        query_time = (time.perf_counter() - start_time) * 1000
        body = (   # 构建查询结果主体信息
            f"找到 {len(ids)} 个{'最近邻点' if mode == 'knn' else '点'} "
            f"(查询耗时: {query_time:.2f}ms)\n\n"
        )

        rows = []   # 遍历查询结果ID
        for i, idx in enumerate(ids, 1):
            if idx >= len(self.spatial_coords):      # 防止索引越界
                continue
            coord = self.spatial_coords[idx]   # 获取对应坐标和数据行
            row = self.spatial_data[idx]
            dist = np.sqrt((coord[0] - lon) ** 2 + (coord[1] - lat) ** 2 + (coord[2] - alt) ** 2)    # 计算三维欧氏距离
            rows.append(    # 构建结果行字符串
                f"{i}. 距离: {dist:.8f}\n"
                f"文件名: {row.get('文件名', 'N/A')}\n"
                f"经度: {coord[0]:.9f}\n"
                f"纬度: {coord[1]:.9f}\n"
                f"高度: {coord[2]:.3f}\n"
            )
        # 清空文本框并插入查询结果
        self.query_result_text.delete(1.0, tk.END)
        self.query_result_text.insert(tk.END, header + body + "\n".join(rows) + ("\n" if rows else ""))


if __name__ == "__main__":
    root = tk.Tk()   # 创建Tkinter根窗口
    UAVExifTool(root)    # 实例化无人机Exif工具主界面
    root.mainloop()   # 进入GUI主事件循环
